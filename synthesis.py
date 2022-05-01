import os
import json
import csv
import time
import datetime
import openpyxl
import sys

class Stock(object):
    def __init__(self,stock_number,source_folder,target_folder,start_date_str,close_date,period):
        self.number = stock_number

        self.data = self.read_all_file(source_folder,start_date_str)

        for time in period:
            self.build_technical_data_and_write(source_folder,target_folder,start_date_str,close_date,time)
        

    def read_all_file(self, source_folder, start_date_str):
        files = os.listdir(source_folder)
        start_date = datetime.datetime.strptime(start_date_str, '%Y%m%d')

        data = []
        for date_str in files:
            date = datetime.datetime.strptime(date_str, '%Y%m%d')
            if date >= start_date and date_str not in close_date:
                filename = '{}/{}/{}.xlsx'.format(source_folder, date_str, self.number)
                new_data = read_excel_file(filename, 'Sheet')
                for row in new_data:
                    tmp_dict = row
                    tmp_dict['Date'] = date_str
                    data.append(tmp_dict)
        return data

    def build_technical_data_and_write(self,source_folder,target_folder,start_date_str,close_date,time):
        last_file = '{}/{}/{}-{}min.xlsx'.format(source_folder, last_date(start_date_str, close_date), self.number, time)
        last_data = read_excel_file(last_file, 'minute data')

        self.KD_num = 9
        self.RSI_num = 14

        self.KD_queue = my_Queue(self.KD_num)
        self.MA5_queue = my_Queue(5)
        self.MA10_queue = my_Queue(10)
        self.MA20_queue = my_Queue(20)

        self.RSI_up_queue = my_Queue(self.RSI_num)
        self.RSI_down_queue = my_Queue(self.RSI_num)
        self.RSI_last_price = 1

        self.sum_price = 0

        for row in last_data:
            time_str = str(row['資料時間'])
            tmp = time_str.split(':')
            minutes = int(tmp[1])
            if minutes % time == 0 :
                price = float(row['當前成交價'])
                self.KD_queue.push(price)
                self.MA5_queue.push(price)
                self.MA10_queue.push(price)
                self.MA20_queue.push(price)
                #self.RSI_dis = (price - self.RSI_last_price) / self.RSI_last_price
                self.RSI_dis = (price - self.RSI_last_price)
                if self.RSI_dis > 0:
                    self.RSI_up_queue.push(self.RSI_dis)
                    self.RSI_down_queue.push(0)
                elif self.RSI_dis < 0:
                    self.RSI_up_queue.push(0)
                    self.RSI_down_queue.push(self.RSI_dis * -1)
                else:
                    self.RSI_up_queue.push(0)
                    self.RSI_down_queue.push(0)

                self.RSI_last_price = price
                self.last_K = row['K']#
                self.last_D = row['D']#
                self.last_EMA12 = row['EMA12']#
                self.last_EMA26 = row['EMA26']#
                self.last_MACD = row['MACD']#
        
        date_str = last_date(start_date_str, close_date)
        last_sum_vol = 0
        array = []
        
        for minute in self.data:
            time_str = str(minute['資料時間'])
            tmp = time_str.split(':')
            minutes = int(tmp[1])

            if minutes % time == 0 :
                if date_str != minute['Date']:
                    self.sum_price = 0
                    last_sum_vol = 0
                    date_str = minute['Date']

                price = float(minute['當前成交價'])
                vol = float(minute['累積成交量']) - last_sum_vol
                sum_vol = float(minute['累積成交量'])
                last_sum_vol = sum_vol
                
                self.cal_RSI(price)
                self.cal_KD(price)
                self.cal_MA(price)
                self.cal_MACD(price)
                self.cal_avg_price(price,vol,sum_vol)
                array.append({
                    'Date' : minute['Date'],
                    'Time' : minute['資料時間'],
                    'Current Price' : minute['當前成交價'],
                    'Volume' : vol,
                    'Total Volume' : minute['累積成交量'],
                    'RSI' : self.RSI,
                    'K' : self.K,
                    'D' : self.D,
                    'MA5' : self.MA5,
                    'MA10' : self.MA10,
                    'MA20' : self.MA20,
                    'MACD' : self.MACD,
                    'OSC' : self.OSC,
                    'DIF' : self.DIF,
                    'EMA12' : self.EMA12,
                    'EMA26' : self.EMA26,
                    'avg price' : self.avg_price
                })

        fieldname = [
            'Date',
            'Time',
            'Current Price',
            'Volume',
            'RSI',
            'K',
            'D',
            'MA5',
            'MA10',
            'MA20',
            'MACD',
            'OSC',
            'DIF',
            'EMA12',
            'EMA26',
            'avg price',
        ]
        target_file = '{}/{}-{}min.csv'.format(target_folder, self.number, time)
        write_csv_file(target_file, fieldname, array)

        return array

    def cal_RSI(self,price):
        #self.RSI_dis = (price - self.RSI_last_price) / self.RSI_last_price
        self.RSI_dis = (price - self.RSI_last_price)
        if self.RSI_dis > 0:
            self.RSI_up_queue.push(self.RSI_dis)
            self.RSI_down_queue.push(0)
        elif self.RSI_dis < 0:
            self.RSI_up_queue.push(0)
            self.RSI_down_queue.push(self.RSI_dis * -1)
        else:
            self.RSI_up_queue.push(0)
            self.RSI_down_queue.push(0)
        
        up_avg = self.RSI_up_queue.sum() / self.RSI_num
        down_avg = self.RSI_down_queue.sum() / self.RSI_num

        if (up_avg + down_avg) == 0:
            self.RSI = 0
        else:
            self.RSI = float(int(up_avg / (up_avg + down_avg)*1000))/10
        
        self.RSI_last_price = price

    def cal_KD(self,cur_price):

        self.KD_queue.push(cur_price)
        
        if self.KD_queue.max() == self.KD_queue.min():
            self.RSV = 0
        else:
            self.RSV = ((cur_price - self.KD_queue.min())*100)/(self.KD_queue.max() - self.KD_queue.min())
        self.K = ((self.last_K)*2 + (self.RSV))/3
        self.D = ((self.last_D)*2 + (self.K))/3
        self.last_K = self.K
        self.last_D = self.D

    def cal_MA(self,price):
        self.MA5_queue.push(price)
        self.MA10_queue.push(price)
        self.MA20_queue.push(price)
        self.MA5 = self.MA5_queue.sum()/5
        self.MA10 = self.MA10_queue.sum()/10
        self.MA20 = self.MA20_queue.sum()/20
    
    def cal_MACD(self,price):
        n = 12
        m = 26
        x = 9

        self.EMA12 = (self.last_EMA12 * (n-1) + price * 2) / (n+1)
        self.EMA26 = (self.last_EMA26 * (m-1) + price * 2) / (m+1)
        self.DIF = self.EMA12 - self.EMA26
        self.MACD = (self.last_MACD * (x-1) + self.DIF * 2) / (x+1)
        self.OSC = self.DIF - self.MACD

        self.last_EMA12 = self.EMA12
        self.last_EMA26 = self.EMA26
        self.last_MACD = self.MACD

    def cal_avg_price(self,price,vol,sum_vol):
        self.sum_price = self.sum_price + (price * vol)
        if sum_vol == 0:
            self.avg_price = 0
        else:
            self.avg_price = self.sum_price / sum_vol

class my_Queue(object):
    def __init__(self,length):
        self.arr = []
        self.size = 0
        self.len = length
    
    def push(self,num):
        if self.size == self.len:
            self.pop()
        self.arr.append(num)
        self.size+=1

    def pop(self):
        if self.size > 0:
            del self.arr[0]
            self.size-=1
    
    def max(self):
        maximum = 0
        for price in self.arr:
            if float(price) > maximum:
                maximum = float(price)
        return maximum

    def min(self):
        minimum = float('inf')
        for price in self.arr:
            if float(price) < minimum:
                minimum = float(price)
        return minimum

    def sum(self):
        sum = 0
        for price in self.arr:
            sum = sum + float(price)
        return sum

    def print(self):
        for price in self.arr:
            print(price)
        print()

def file_to_array(file):
    workbook = openpyxl.load_workbook(file)
    sheets = workbook.sheetnames
    booksheet = workbook[sheets[0]] 
    rows = booksheet.rows

    array = []
    for row in rows:
        line=[col.value for col in row]
        array.append(str(line[0]))

    return array

def last_date(str_today, close):
    today = datetime.datetime.strptime(str_today, '%Y%m%d')
    oneday = datetime.timedelta(days=1) 
    
    weekday = today.weekday() + 1
    if weekday == 1:
        last_date = today - (oneday * 3)
    else:
        last_date = today - oneday
    str_last = last_date.strftime('%Y%m%d')

    while str_last in close:
        last_date = datetime.datetime.strptime(str_last, '%Y%m%d')
        weekday = last_date.weekday() + 1
        if weekday == 1:
            last_date = last_date - (oneday * 3)
        else:
            last_date = last_date - oneday
        str_last = last_date.strftime('%Y%m%d')
    
    return str_last

def write_csv_file(target_file, fieldnames, data):
    with open(target_file, 'w', newline='') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames = fieldnames, extrasaction='ignore')
        writer.writeheader()
        
        for row in data:
            writer.writerow(row)

def read_excel_file(original,sheet):
    workbook = openpyxl.load_workbook(original)

    booksheet = workbook[sheet] 
    rows = booksheet.rows
    array = []
    ref = []
    
    for row in rows:
        line = [col.value for col in row]
        if len(line)!=0 and line[0]!='資料時間' and line[0]!='tv' and line[0] is not None:
            line[0]=str(line[0])
            tmp = {i : j for i,j in zip(ref,line)}
            array.append(tmp)
        else:
            ref = line
    return array

source_folder = 'data'
target_folder = 'synthesis data'
if not os.path.isdir(target_folder):
    os.mkdir(target_folder)

start_date_str = '20210107'

stocks = file_to_array('stocknumber.xlsx')
close_date = file_to_array('close.xlsx')

period = [1,5,10]

for stock in stocks:
    Stock(stock,source_folder,target_folder,start_date_str,close_date,period)

