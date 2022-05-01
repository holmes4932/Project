#!/usr/bin/python
# -*- coding: utf-8 -*-
import datetime
import openpyxl
import os
'''
"result_of_buy_股票名稱"的記事本 : 裡面的worksheet有total、日期
total : 統整每天的報酬率和賺多少錢，並算出日期的平均報酬率和賺多少錢
日期 : 儲存每天的交易資料(買賣點、一天的平均報酬率、賺多少)

程式開始會去讀"result_of_buy_股票名稱.xlsx"
新增當天的日期worksheet，儲存交易資料
最後會讀取total，將當天的資料加上去，並計算總平均
'''

class Strategy():
    def __init__(self, stocknumber, date, money):
        self.stocknumber = stocknumber
        self.money = int(money) #本金
        self.left_money = money #剩餘的本金
        self.date = date

        self.sum = 0
        self.num = 0 #身上總共有幾張股票
        self.ori = 0 #目前價格
        self.avg = 0
        self.buy_sum = 0
        self.init = 0 #開盤價
        self.view = []

        #初始化workbook
        self.w = 1 #計算放到worksheet的第幾行
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.total_worksheet = self.workbook.active
        self.init_workbook()

    #每分鐘push一次
    def push(self, data):
        time = data['time']
        ori = data['original']
        one = data['1min']
        five = data['5min']
        ten = data['10min']
        volume = data['volume']
        ma5 = data['MA5']
        avg_price = data['avg_price']
        self.ori = ori

        #紀錄開盤價
        start_time = datetime.datetime.strptime(str(time.date()) + ' 09:00', '%Y-%m-%d %H:%M')
        if(time == start_time) :
            self.init = ori

        lot = self.decision2(time, ori, one, five, ten, self.num, self.avg, volume, ma5, avg_price, self.left_money, self.init)
        
        #在13:20以後就將left歸零不再買進
        stop_buying_time = datetime.datetime.strptime(str(time.date()) + ' 13:20', '%Y-%m-%d %H:%M')
        if(time >= stop_buying_time) :
            lot = 0
        #print('買進', lot, '張')
        self.num = self.num + lot

        #沒有用到這個avg和volume，有空可以刪掉
        #買入的平均成本
        if(self.num == 0) :
            self.buy_sum = 0
            self.avg = 0
        else :
            self.buy_sum = self.sum + lot * ori + abs(lot * ori * 0.001425)
            self.avg = self.buy_sum / self.num   
        
 
        #總共花多少錢
        if(lot > 0) :
            self.sum = self.sum + lot * ori + lot * ori * 0.001425
        elif(lot < 0) :
            self.sum = self.sum + lot * ori + abs(lot * ori * 0.002925)

        #本金 - 花的錢 -> 計算剩下多少錢可以買    
        self.left_money = self.money - self.sum * 1000

        #write worksheet
        lot = self.write_worksheet(time, lot, ori)
        #回傳交易的資料
        line = {}
        hm = time.strftime('%H:%M')
        if(lot > 0) :
            line['時間'] = hm
            line['買入張數'] = lot
            line['買入價格'] = ori
            line['賣出張數'] = '-'
            line['賣出價格'] = '-'
            line['持有張數'] = self.num
            self.view.append(line)
        elif (lot < 0) :
            line['時間'] = hm
            line['買入張數'] = '-'
            line['買入價格'] = '-'
            line['賣出張數'] = -lot
            line['賣出價格'] = ori
            line['持有張數'] = self.num
            self.view.append(line)


    #最後印出
    def overview(self):
        #報酬率
        return_rate = -self.sum * 1000 * 100 / self.money
        print('賺了:',"{:.3f}".format(-self.sum * 1000),'元')
        print('報酬率:', "{:.3f}".format(return_rate), '%')
        self.final_write_worksheet(return_rate)
        #return -self.sum, return_rate

    #用MA5跟均價做止損
    def decision(self, time, ori, one, five, ten, num, avg, volume, MA5, avg_price, left, init) :
        m1 = one - ori
        m2 = five - ori
        m3 = ten - ori
        fee = ori * 0.001425
        big_fee = ori* 0.00435
        #漲停限制&跌停限制 : -8% < 股價 < 8% 才可買入
        #1、5min漲幅皆大於買入手續費(0.001425) 10 min > 買賣手續費(0.00435) 且 10 > 5 > 1min預測價格
        if (m3 > big_fee and m2 > fee and m1 > fee and m3 > m2 > m1 and ori > init*0.92 and ori < init*1.08) :
            if(left - ori*1.001425*1000 < 0) :
                print("not enough !! you have left %f" %left)
                return 0
            else : 
                print( "time is %s" %time)
                print("left %f" %left)
                print( "buy in %f" %ori)
                return 1
        #漲停限制&跌停限制 : 股價 > 9% 或 股價 < -9% 要賣掉
        elif ((ori < init*1.09 or ori > init*0.91)and num > 0) : 
            num = 0 - num
            return num
        #1、5、10min預測價格漲幅皆為負
        elif (m3 < 0 and m2 < 0 and m1 < 0 and num > 0) :
            num = 0 - num
            print( "time is %s" %time)
            print( "預期不會漲sold in %f" %ori)
            return num
        #1、5、10min預測價格皆小於MA
        elif (MA5 > one and MA5 > five and MA5 > ten and num > 0) :
            num = 0 - num
            print( "time is %s" %time)
            print( "MA停損 sold in %f" %ori)
            return num
        #1、5、10min預測價格皆小於均價
        elif (avg_price > one and avg_price > five and avg_price > ten and num > 0) :
            num = 0 - num
            print( "time is %s" %time)
            print( "均價停損 sold in %f" %ori)
            return num

        else :
            return 0

    #不做止損
    def decision2(self, time, ori, one, five, ten, num, avg, volume, MA5, avg_price, left, init) :
        m1 = one - ori
        m2 = five - ori
        m3 = ten - ori
        fee = ori * 0.001425
        big_fee = ori* 0.00435
        #漲停限制&跌停限制 : -8% < 股價 < 8% 才可買入
        #1、5min漲幅皆大於買入手續費(0.001425) 10 min > 買賣手續費(0.00435) 且 10 > 5 > 1min預測價格
        if (m3 > big_fee and m2 > fee and m1 > fee and m3 > m2 > m1 and ori > init*0.92 and ori < init*1.08) :
            if(left - ori*1.001425*1000 < 0) :
                print("not enough !! you have left %f" %left)
                return 0
            else : 
                print( "time is %s" %time)
                print("left %f" %left)
                print( "buy in %f" %ori)
                return 1
        #漲停限制&跌停限制 : 股價 > 9% 或 股價 < -9% 要賣掉
        elif ((ori < init*0.91 or ori > init*1.09)and num > 0) : 
            num = 0 - num
            return num
        #1、5、10min預測價格漲幅皆為負
        elif (m3 < 0 and m2 < 0 and m1 < 0 and num > 0) :
            num = 0 - num
            print( "time is %s" %time)
            print( "預期不會漲sold in %f" %ori)
            return num

        else :
            return 0

    #只用MA5做止損
    def decision3(self, time, ori, one, five, ten, num, avg, volume, MA5, avg_price, left, init) :
        m1 = one - ori
        m2 = five - ori
        m3 = ten - ori
        fee = ori * 0.001425
        big_fee = ori* 0.00435
        #漲停限制&跌停限制 : -8% < 股價 < 8% 才可買入
        #1、5min漲幅皆大於買入手續費(0.001425) 10 min > 買賣手續費(0.00435) 且 10 > 5 > 1min預測價格
        if (m3 > big_fee and m2 > fee and m1 > fee and m3 > m2 > m1 and ori > init*0.92 and ori < init*1.08) :
            if(left - ori*1.001425*1000 < 0) :
                print("not enough !! you have left %f" %left)
                return 0
            else : 
                print( "time is %s" %time)
                print("left %f" %left)
                print( "buy in %f" %ori)
                return 1
        #漲停限制&跌停限制 : 股價 > 9% 或 股價 < -9% 要賣掉
        elif ((ori < init*1.09 or ori > init*0.91)and num > 0) : 
            num = 0 - num
            return num
        #1、5、10min預測價格漲幅皆為負
        elif (m3 < 0 and m2 < 0 and m1 < 0 and num > 0) :
            num = 0 - num
            print( "time is %s" %time)
            print( "預期不會漲sold in %f" %ori)
            return num
        #1、5、10min預測價格皆小於MA
        elif (MA5 > one and MA5 > five and MA5 > ten and num > 0) :
            num = 0 - num
            print( "time is %s" %time)
            print( "MA停損 sold in %f" %ori)
            return num
        else :
            return 0

    #只用均價做止損
    def decision4(self, time, ori, one, five, ten, num, avg, volume, MA5, avg_price, left, init) :
        m1 = one - ori
        m2 = five - ori
        m3 = ten - ori
        fee = ori * 0.001425
        big_fee = ori* 0.00435
        #漲停限制&跌停限制 : -8% < 股價 < 8% 才可買入
        #1、5min漲幅皆大於買入手續費(0.001425) 10 min > 買賣手續費(0.00435) 且 10 > 5 > 1min預測價格
        if (m3 > big_fee and m2 > fee and m1 > fee and m3 > m2 > m1 and ori > init*0.92 and ori < init*1.08) :
            if(left - ori*1.001425*1000 < 0) :
                print("not enough !! you have left %f" %left)
                return 0
            else : 
                print( "time is %s" %time)
                print("left %f" %left)
                print( "buy in %f" %ori)
                return 1
        #漲停限制&跌停限制 : 股價 > 9% 或 股價 < -9% 要賣掉
        elif ((ori < init*1.09 or ori > init*0.91) and num > 0) : 
            num = 0 - num
            return num
        #1、5、10min預測價格漲幅皆為負
        elif (m3 < 0 and m2 < 0 and m1 < 0 and num > 0) :
            num = 0 - num
            print( "time is %s" %time)
            print( "預期不會漲sold in %f" %ori)
            return num
        #1、5、10min預測價格皆小於均價
        elif (avg_price > one and avg_price > five and avg_price > ten and num > 0) :
            num = 0 - num
            print( "time is %s" %time)
            print( "均價停損 sold in %f" %ori)
            return num

        else :
            return 0

    def get_left_money(self): 
        return int(self.left_money) 
 
    def get_lot_num(self): 
        return self.num 
     
    def get_return_rate(self): 
        now = -self.sum * 1000 + self.ori*self.num * 1000
        rate = now*100 / self.money 
        rate = round(rate, 3)
        return rate

    def get_detail_view(self) :
        return self.view
        
    def init_workbook(self) :
        #write worksheet
        path = 'final result'
        filepath = 'final result/result_of_buy_%s.xlsx' %self.stocknumber
        if not os.path.isdir(path):
            os.mkdir(path)
        if not os.path.isfile(filepath) :
            self.total_worksheet = self.workbook['Sheet']
            self.total_worksheet.title = "total"
            self.total_worksheet.cell(row = 1, column = 1, value ='day')
            self.total_worksheet.cell(row = 1, column = 2, value ='return rate')
            self.total_worksheet.cell(row = 1, column = 3, value ='earn(元)')
        else :
            self.workbook = openpyxl.load_workbook('final result/result_of_buy_%s.xlsx' %self.stocknumber)
            self.total_worksheet = self.workbook['total']
        #判斷worksheet的sheetname是否已存在，存在的話刪除worksheet
        if self.date in self.workbook.sheetnames :
            re_sheet = self.workbook[self.date]
            self.workbook.remove(re_sheet)
        self.worksheet = self.workbook.create_sheet('%s' %self.date)
        self.worksheet.cell(row = 1, column = 1, value ='time')
        self.worksheet.column_dimensions['A'].width = 18.0
        self.worksheet.cell(row = 1, column = 2, value ='buy number')
        self.worksheet.cell(row = 1, column = 3, value ='buy price')
        self.worksheet.cell(row = 1, column = 4, value ='sell number')
        self.worksheet.cell(row = 1, column = 5, value ='sell price')
        self.worksheet.cell(row = 1, column = 6, value ='total number')
        self.w = 2 

    def write_worksheet(self, time, lot, ori) :
        if(lot > 0 ) :
            self.worksheet.cell(row = self.w, column = 1, value = time)
            self.worksheet.cell(row = self.w, column = 2, value = lot)
            self.worksheet.cell(row = self.w, column = 3, value = ori)
            self.worksheet.cell(row = self.w, column = 6, value = self.num)
            self.w = self.w + 1
        elif (lot < 0) :
            self.worksheet.cell(row = self.w, column = 1, value = time)
            self.worksheet.cell(row = self.w, column = 4, value = -lot)
            self.worksheet.cell(row = self.w, column = 5, value = ori)
            self.worksheet.cell(row = self.w, column = 6, value = self.num)
            self.w = self.w + 1

        #最後一分鐘售出剩下的股票 
        last_min = datetime.datetime.strptime(str(time.date()) + ' 13:29', '%Y-%m-%d %H:%M')
        if(time >= last_min and self.num > 0) :  
            self.sum = self.sum + self.num * (-ori) + abs( self.num * ori * 0.002925)
            #write worksheet
            self.worksheet.cell(row = self.w, column = 1, value = time)
            self.worksheet.cell(row = self.w, column = 4, value = self.num)
            self.worksheet.cell(row = self.w, column = 5, value = ori)
            self.worksheet.cell(row = self.w, column = 6, value = 0)
            self.w = self.w + 1
            lot = -self.num
            self.num = 0
        return lot

    def final_write_worksheet(self, return_rate) :
        #write worksheet
        self.worksheet.cell(row = self.w+1, column = 1, value = '本金(元) : %s' %self.money)
        self.worksheet.cell(row = self.w+2, column = 1, value = '賺了(元) : %s' %"{:.3f}".format(-self.sum*1000))
        self.worksheet.cell(row = self.w+3, column = 1, value = '報酬率(%%): %s' %"{:.3f}".format(return_rate))
        #write total worksheet
        total_return_rate = 0
        #計算總報酬率，並檢查是否為重複的日期，是的話覆蓋
        c = 2 #有幾筆資料
        f = 0 # f=1時，表示有重複的日期
        while self.total_worksheet.cell(row = c , column = 1).value :
            if self.total_worksheet.cell(row = c , column = 1).value == self.date :
                total_return_rate = total_return_rate + return_rate
                self.total_worksheet.cell(row = c, column = 1, value = self.date)
                self.total_worksheet.cell(row = c, column = 2, value = "{:.3f}".format(return_rate))
                self.total_worksheet.cell(row = c, column = 3, value = "{:.3f}".format(-self.sum*1000))
                f = 1
            else :
                total_return_rate = total_return_rate + float(self.total_worksheet.cell(row = c, column = 2).value)
            c = c + 1
        if(f == 0) :
            total_return_rate = total_return_rate + return_rate
            self.total_worksheet.cell(row = c, column = 1, value = self.date)
            self.total_worksheet.cell(row = c, column = 2, value = "{:.3f}".format(return_rate))
            self.total_worksheet.cell(row = c, column = 3, value = "{:.3f}".format(-self.sum*1000))
        #重複資料，因此資料筆數-1
        else :
            c = c - 1
        avg_return_rate = total_return_rate / (c-1)
        self.total_worksheet.cell(row = 1, column = 5, value = '---平均---' )
        self.total_worksheet.cell(row = 2, column = 5, value = '本金(元) : %s' %self.money)
        self.total_worksheet.cell(row = 3, column = 5, value = '賺了(元) : %s' %"{:.3f}".format(avg_return_rate*self.money/100))
        self.total_worksheet.cell(row = 4, column = 5, value = '報酬率(%%): %s' %"{:.3f}".format(avg_return_rate))
        self.workbook.save('final result/result_of_buy_%s.xlsx' %self.stocknumber)
'''
s = Strategy('00637L', 30000)

time = datetime.datetime.strptime('2021/9/7 09:00', '%Y/%m/%d %H:%M')

data = {
    'time':time,
    'original':234.5,
    '1min':234.0164,
    '5min':234.8789,
    '10min':236.4333,
    'volume':4,
    'MA5':232.9,
    'avg_price':234.5
}

s.push(data)
s.overview()
'''
